import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

#Data Siswa
data = {
    "nama": ["budi", "andi", "siti", "rian"],
    "kelas": ["X", "X", "XI", "XI"],
    "nilai" : [80, 90, 75, 85]
}

df = pd.DataFrame(data)

# Buat Workbook baru
wb = Workbook()

#....Sheet 1: Data Siswa....
ws1 = wb.active
ws1.title = "Data Siswa"

for r in dataframe_to_rows(df, index=False, header=True):
    ws1.append (r)

    #....Sheet 2: Anaisis....
    ws2 = wb.create_sheet("analisis")

    rata_rata = df["nilai"].mean()
    nilai_tertinggi = df["nilai"].max()
    nilai_terendah = df["nilai"].min()

    analisis = [
        ["Kategori", "Nilai"],
        ["Rata-rata", rata_rata],
        ["Nilai Tertinggi", nilai_tertinggi],
        ["Nilai Terendah", nilai_terendah]
    ]

    for row in analisis :
        ws2.append(row)

#....Sheet 3: Grafik...
ws3 = wb.create_sheet("grafik")

#Masukkan data untuk chart
for r in dataframe_to_rows(df, index=False, header=True):
    ws3.append(r)

#Buat chart batang
chart = BarChart()
chart.title = "Perbandingan Nilai Siswa"
chart.x_axis.title = "Nama"
chart.y_axis.title = "Nilai"

data_ref = Reference(ws3, min_col=3, min_row=1, max_row=len(df)+1) #Kolom Nilai
cats_ref = Reference(ws3, min_col=1,min_row=2, max_row=len(df)+1) #Kolom Nama

chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)

ws3.add_chart(chart, "E5")

#Simpan file Excel
wb.save("hasil_analisis.xlsx")

print("Excel berhasil di buat: hasil_analisis.xlsx")